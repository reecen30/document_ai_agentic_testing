"""
create_workbooks.py

Run this script once to set up the three Excel workbooks:
  1. DocumentAI_EvidenceStore.xlsx  — source evidence (with mock data)
  2. AgenticTesting_AuditLog.xlsx   — empty audit log (pre-structured sheets)
  3. Run_TEMPLATE_Report.xlsx       — empty run report template

Usage:
    python scripts/create_workbooks.py [--output-dir ./data]
"""

import argparse
import os
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _write_sheet(wb: Workbook, name: str, headers: list, rows: list = None) -> None:
    """
    Create a worksheet with styled headers and optional data rows.

    - Headers: row 1, blue background (#1F4E79), white bold font, centered.
    - Auto-fits column widths based on the max content length in each column.
    """
    ws = wb.create_sheet(title=name)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Write data rows
    if rows:
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths
    col_widths = [len(str(h)) for h in headers]
    if rows:
        for row_data in rows:
            for col_idx, value in enumerate(row_data):
                col_widths[col_idx] = max(col_widths[col_idx], len(str(value)) if value is not None else 0)

    for col_idx, width in enumerate(col_widths, start=1):
        # Add a small buffer and cap at 60
        adjusted = min(width + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Freeze the header row
    ws.freeze_panes = ws["A2"]


# ---------------------------------------------------------------------------
# Workbook 1 — DocumentAI_EvidenceStore.xlsx
# ---------------------------------------------------------------------------

def _create_evidence_store(output_dir: str) -> str:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- DocumentTypes ---
    _write_sheet(
        wb,
        "DocumentTypes",
        headers=["DocumentTypeID", "DocumentType"],
        rows=[
            (1, "ApplicationForm"),
            (2, "Resolution"),
            (3, "Windeed"),
            (4, "Passport"),
            (5, "IdentityDocument"),
            (6, "ProductFormsBTB"),
            (7, "ProductFormsIFB"),
            (8, "ProductFormsICIB"),
            (9, "Income_Statement"),
            (10, "Balance_Sheet"),
            (11, "CashFlow"),
            (12, "Debtors"),
            (13, "Creditors"),
            (14, "AFS"),
            (15, "Other"),
            (16, "RelatedPartyForm"),
        ],
    )

    # --- ProcessStages ---
    _write_sheet(
        wb,
        "ProcessStages",
        headers=["ProcessStageID", "ProcessStageName"],
        rows=[
            (1, "Pre Classify"),
            (2, "Validated Post Classified"),
            (3, "Pre Extract"),
            (4, "Validated Post Extracted"),
        ],
    )

    # --- ai.ModelStages ---
    _write_sheet(
        wb,
        "ai.ModelStages",
        headers=["ModelStageID", "ModelStageName"],
        rows=[
            (1, "Input"),
            (2, "Output"),
        ],
    )

    # --- ai.ModelTypes ---
    _write_sheet(
        wb,
        "ai.ModelTypes",
        headers=["ModelTypeID", "ModelTypeName"],
        rows=[
            (1, "LLM"),
            (2, "Agent"),
        ],
    )

    # --- ai.ModelNames ---
    _write_sheet(
        wb,
        "ai.ModelNames",
        headers=["ModelNameID", "ModelName"],
        rows=[
            (9001, "Agentic_IntakeDiff"),
            (9002, "Agentic_ScopePlanner"),
            (9003, "Agentic_EvidenceCollector"),
            (9004, "Agentic_RegressionHunter"),
            (9005, "Agentic_Challenger"),
            (9006, "Agentic_TargetedRerun"),
            (9007, "Agentic_TrendDrift"),
            (9008, "Agentic_RootCause"),
            (9009, "Agentic_PatchProposal"),
            (9010, "Agentic_ReportRouting"),
        ],
    )

    # --- DocumentData ---
    doc_data_headers = [
        "DocumentDataID", "TransactionID", "DocumentID", "ProcessStageID",
        "DocumentTypeID", "DocumentTypeName", "StartPage", "Confidence",
        "OcrConfidence", "Field", "IsMissing", "Value",
        "FieldConfidence", "FieldOcrConfidence", "Created_By",
        "CreatedDateTime", "Active",
    ]

    doc_data_rows = [
        # ---------------------------------------------------------------
        # Scenario 1: TXN 1001 — stable gold (IdentityDocument correct)
        # ---------------------------------------------------------------
        (1, 1001, "1001_pack", 1, 5, "IdentityDocument", 1, 0.91, 0.98,
         "DocumentType", 0, "IdentityDocument", 0.91, 0.98, "system",
         "2026-03-14 10:00:00", 1),
        (2, 1001, "1001_pack", 2, 5, "IdentityDocument", 1, 1.00, 0.98,
         "DocumentType", 0, "IdentityDocument", 1.00, 0.98, "validator",
         "2026-03-14 10:05:00", 1),
        (3, 1001, "1001_pack", 3, 5, "IdentityDocument", 1, 0.89, 0.98,
         "IdentityNumber", 0, "8001015009087", 0.89, 0.98, "system",
         "2026-03-14 10:00:00", 1),
        (4, 1001, "1001_pack", 4, 5, "IdentityDocument", 1, 1.00, 0.98,
         "IdentityNumber", 0, "8001015009087", 1.00, 0.98, "validator",
         "2026-03-14 10:05:00", 1),

        # ---------------------------------------------------------------
        # Scenario 2: TXN 1002 — ID/Passport confusion
        # Baseline classifies as Passport; truth is IdentityDocument
        # ---------------------------------------------------------------
        (5, 1002, "1002_pack", 1, 4, "Passport", 1, 0.83, 0.97,
         "DocumentType", 0, "Passport", 0.83, 0.97, "system",
         "2026-03-14 10:00:00", 1),
        (6, 1002, "1002_pack", 2, 5, "IdentityDocument", 1, 1.00, 0.97,
         "DocumentType", 0, "IdentityDocument", 1.00, 0.97, "validator",
         "2026-03-14 10:05:00", 1),
        (7, 1002, "1002_pack", 3, 4, "Passport", 1, 0.72, 0.97,
         "PassportNumber", 0, "A1234567", 0.72, 0.97, "system",
         "2026-03-14 10:00:00", 1),
        (8, 1002, "1002_pack", 4, 5, "IdentityDocument", 1, 1.00, 0.97,
         "IdentityNumber", 0, "9002026009088", 1.00, 0.97, "validator",
         "2026-03-14 10:05:00", 1),

        # ---------------------------------------------------------------
        # Scenario 3: TXN 1003 — out-of-scope (both agree: Other)
        # ---------------------------------------------------------------
        (9, 1003, "1003_pack", 1, 15, "Other", 1, 0.88, 0.95,
         "DocumentType", 0, "Other", 0.88, 0.95, "system",
         "2026-03-14 10:00:00", 1),
        (10, 1003, "1003_pack", 2, 15, "Other", 1, 1.00, 0.95,
         "DocumentType", 0, "Other", 1.00, 0.95, "validator",
         "2026-03-14 10:05:00", 1),
        (11, 1003, "1003_pack", 3, 15, "Other", 1, 0.81, 0.95,
         "Comment", 0, "Out of scope brochure", 0.81, 0.95, "system",
         "2026-03-14 10:00:00", 1),
        (12, 1003, "1003_pack", 4, 15, "Other", 1, 1.00, 0.95,
         "Comment", 0, "Out of scope brochure", 1.00, 0.95, "validator",
         "2026-03-14 10:05:00", 1),

        # ---------------------------------------------------------------
        # Scenario 4: TXN 1004 — extraction completeness issue
        # Baseline has a missing field (EntityName)
        # ---------------------------------------------------------------
        (13, 1004, "1004_pack", 1, 1, "ApplicationForm", 1, 0.92, 0.96,
         "DocumentType", 0, "ApplicationForm", 0.92, 0.96, "system",
         "2026-03-14 10:00:00", 1),
        (14, 1004, "1004_pack", 2, 1, "ApplicationForm", 1, 1.00, 0.96,
         "DocumentType", 0, "ApplicationForm", 1.00, 0.96, "validator",
         "2026-03-14 10:05:00", 1),
        (15, 1004, "1004_pack", 3, 1, "ApplicationForm", 1, 0.00, 0.96,
         "EntityName", 1, "", 0.00, 0.96, "system",
         "2026-03-14 10:00:00", 1),
        (16, 1004, "1004_pack", 4, 1, "ApplicationForm", 1, 1.00, 0.96,
         "EntityName", 0, "Acme Holdings Ltd", 1.00, 0.96, "validator",
         "2026-03-14 10:05:00", 1),

        # ---------------------------------------------------------------
        # Scenario 5: TXN 1005 — high confidence wrong answer
        # Baseline confident but wrong (Resolution vs ApplicationForm)
        # ---------------------------------------------------------------
        (17, 1005, "1005_pack", 1, 2, "Resolution", 1, 0.94, 0.97,
         "DocumentType", 0, "Resolution", 0.94, 0.97, "system",
         "2026-03-14 10:00:00", 1),
        (18, 1005, "1005_pack", 2, 1, "ApplicationForm", 1, 1.00, 0.97,
         "DocumentType", 0, "ApplicationForm", 1.00, 0.97, "validator",
         "2026-03-14 10:05:00", 1),
        (19, 1005, "1005_pack", 3, 2, "Resolution", 1, 0.91, 0.97,
         "ResolutionDate", 0, "2026-01-15", 0.91, 0.97, "system",
         "2026-03-14 10:00:00", 1),
        (20, 1005, "1005_pack", 4, 1, "ApplicationForm", 1, 1.00, 0.97,
         "ApplicationDate", 0, "2026-01-15", 1.00, 0.97, "validator",
         "2026-03-14 10:05:00", 1),

        # ---------------------------------------------------------------
        # Scenario 6: TXN 1006 - stable gold (IdentityDocument correct)
        # ---------------------------------------------------------------
        (21, 1006, "1006_pack", 1, 5, "IdentityDocument", 1, 0.90, 0.96,
         "DocumentType", 0, "IdentityDocument", 0.90, 0.96, "system",
         "2026-03-14 10:10:00", 1),
        (22, 1006, "1006_pack", 2, 5, "IdentityDocument", 1, 1.00, 0.96,
         "DocumentType", 0, "IdentityDocument", 1.00, 0.96, "validator",
         "2026-03-14 10:15:00", 1),
        (23, 1006, "1006_pack", 3, 5, "IdentityDocument", 1, 0.89, 0.96,
         "IdentityNumber", 0, "8503035009089", 0.89, 0.96, "system",
         "2026-03-14 10:10:00", 1),
        (24, 1006, "1006_pack", 4, 5, "IdentityDocument", 1, 1.00, 0.96,
         "IdentityNumber", 0, "8503035009089", 1.00, 0.96, "validator",
         "2026-03-14 10:15:00", 1),

        # ---------------------------------------------------------------
        # Scenario 7: TXN 1007 - baseline wrong, truth Passport
        # ---------------------------------------------------------------
        (25, 1007, "1007_pack", 1, 15, "Other", 1, 0.82, 0.94,
         "DocumentType", 0, "Other", 0.82, 0.94, "system",
         "2026-03-14 10:10:00", 1),
        (26, 1007, "1007_pack", 2, 4, "Passport", 1, 1.00, 0.94,
         "DocumentType", 0, "Passport", 1.00, 0.94, "validator",
         "2026-03-14 10:15:00", 1),
        (27, 1007, "1007_pack", 3, 15, "Other", 1, 0.78, 0.94,
         "Comment", 0, "Unclear document", 0.78, 0.94, "system",
         "2026-03-14 10:10:00", 1),
        (28, 1007, "1007_pack", 4, 4, "Passport", 1, 1.00, 0.94,
         "PassportNumber", 0, "B7654321", 1.00, 0.94, "validator",
         "2026-03-14 10:15:00", 1),

        # ---------------------------------------------------------------
        # Scenario 8: TXN 1008 - extraction missing in baseline
        # ---------------------------------------------------------------
        (29, 1008, "1008_pack", 1, 1, "ApplicationForm", 1, 0.93, 0.97,
         "DocumentType", 0, "ApplicationForm", 0.93, 0.97, "system",
         "2026-03-14 10:10:00", 1),
        (30, 1008, "1008_pack", 2, 1, "ApplicationForm", 1, 1.00, 0.97,
         "DocumentType", 0, "ApplicationForm", 1.00, 0.97, "validator",
         "2026-03-14 10:15:00", 1),
        (31, 1008, "1008_pack", 3, 1, "ApplicationForm", 1, 0.00, 0.97,
         "ApplicationDate", 1, "", 0.00, 0.97, "system",
         "2026-03-14 10:10:00", 1),
        (32, 1008, "1008_pack", 4, 1, "ApplicationForm", 1, 1.00, 0.97,
         "ApplicationDate", 0, "2026-02-20", 1.00, 0.97, "validator",
         "2026-03-14 10:15:00", 1),

        # ---------------------------------------------------------------
        # Scenario 9: TXN 1009 - stable out-of-scope (Other)
        # ---------------------------------------------------------------
        (33, 1009, "1009_pack", 1, 15, "Other", 1, 0.88, 0.92,
         "DocumentType", 0, "Other", 0.88, 0.92, "system",
         "2026-03-14 10:10:00", 1),
        (34, 1009, "1009_pack", 2, 15, "Other", 1, 1.00, 0.92,
         "DocumentType", 0, "Other", 1.00, 0.92, "validator",
         "2026-03-14 10:15:00", 1),
        (35, 1009, "1009_pack", 3, 15, "Other", 1, 0.79, 0.92,
         "Comment", 0, "Marketing brochure", 0.79, 0.92, "system",
         "2026-03-14 10:10:00", 1),
        (36, 1009, "1009_pack", 4, 15, "Other", 1, 1.00, 0.92,
         "Comment", 0, "Marketing brochure", 1.00, 0.92, "validator",
         "2026-03-14 10:15:00", 1),

        # ---------------------------------------------------------------
        # Scenario 10: TXN 1010 - stable gold (Passport correct)
        # ---------------------------------------------------------------
        (37, 1010, "1010_pack", 1, 4, "Passport", 1, 0.91, 0.97,
         "DocumentType", 0, "Passport", 0.91, 0.97, "system",
         "2026-03-14 10:10:00", 1),
        (38, 1010, "1010_pack", 2, 4, "Passport", 1, 1.00, 0.97,
         "DocumentType", 0, "Passport", 1.00, 0.97, "validator",
         "2026-03-14 10:15:00", 1),
        (39, 1010, "1010_pack", 3, 4, "Passport", 1, 0.90, 0.97,
         "PassportNumber", 0, "P8899001", 0.90, 0.97, "system",
         "2026-03-14 10:10:00", 1),
        (40, 1010, "1010_pack", 4, 4, "Passport", 1, 1.00, 0.97,
         "PassportNumber", 0, "P8899001", 1.00, 0.97, "validator",
         "2026-03-14 10:15:00", 1),

        # ---------------------------------------------------------------
        # Scenario 11: TXN 1011 - onboarding Windeed improvement
        # ---------------------------------------------------------------
        (41, 1011, "1011_pack", 1, 15, "Other", 1, 0.78, 0.93,
         "DocumentType", 0, "Other", 0.78, 0.93, "system",
         "2026-03-14 10:20:00", 1),
        (42, 1011, "1011_pack", 2, 3, "Windeed", 1, 1.00, 0.93,
         "DocumentType", 0, "Windeed", 1.00, 0.93, "validator",
         "2026-03-14 10:25:00", 1),
        (43, 1011, "1011_pack", 3, 15, "Other", 1, 0.72, 0.93,
         "Comment", 0, "Property search printout", 0.72, 0.93, "system",
         "2026-03-14 10:20:00", 1),
        (44, 1011, "1011_pack", 4, 3, "Windeed", 1, 1.00, 0.93,
         "DeedNumber", 0, "WD-556677", 1.00, 0.93, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 12: TXN 1012 - Windeed extraction completeness issue
        # ---------------------------------------------------------------
        (45, 1012, "1012_pack", 1, 3, "Windeed", 1, 0.87, 0.94,
         "DocumentType", 0, "Windeed", 0.87, 0.94, "system",
         "2026-03-14 10:20:00", 1),
        (46, 1012, "1012_pack", 2, 3, "Windeed", 1, 1.00, 0.94,
         "DocumentType", 0, "Windeed", 1.00, 0.94, "validator",
         "2026-03-14 10:25:00", 1),
        (47, 1012, "1012_pack", 3, 3, "Windeed", 1, 0.00, 0.94,
         "OwnerName", 1, "", 0.00, 0.94, "system",
         "2026-03-14 10:20:00", 1),
        (48, 1012, "1012_pack", 4, 3, "Windeed", 1, 1.00, 0.94,
         "OwnerName", 0, "Blue Sky Properties", 1.00, 0.94, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 13: TXN 1013 - ProductFormsBTB stable
        # ---------------------------------------------------------------
        (49, 1013, "1013_pack", 1, 6, "ProductFormsBTB", 1, 0.90, 0.95,
         "DocumentType", 0, "ProductFormsBTB", 0.90, 0.95, "system",
         "2026-03-14 10:20:00", 1),
        (50, 1013, "1013_pack", 2, 6, "ProductFormsBTB", 1, 1.00, 0.95,
         "DocumentType", 0, "ProductFormsBTB", 1.00, 0.95, "validator",
         "2026-03-14 10:25:00", 1),
        (51, 1013, "1013_pack", 3, 6, "ProductFormsBTB", 1, 0.88, 0.95,
         "ProductCode", 0, "PF-BTB-13", 0.88, 0.95, "system",
         "2026-03-14 10:20:00", 1),
        (52, 1013, "1013_pack", 4, 6, "ProductFormsBTB", 1, 1.00, 0.95,
         "ProductCode", 0, "PF-BTB-13", 1.00, 0.95, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 14: TXN 1014 - ProductFormsBTB candidate regression
        # ---------------------------------------------------------------
        (53, 1014, "1014_pack", 1, 6, "ProductFormsBTB", 1, 0.89, 0.94,
         "DocumentType", 0, "ProductFormsBTB", 0.89, 0.94, "system",
         "2026-03-14 10:20:00", 1),
        (54, 1014, "1014_pack", 2, 6, "ProductFormsBTB", 1, 1.00, 0.94,
         "DocumentType", 0, "ProductFormsBTB", 1.00, 0.94, "validator",
         "2026-03-14 10:25:00", 1),
        (55, 1014, "1014_pack", 3, 6, "ProductFormsBTB", 1, 0.86, 0.94,
         "ProductCode", 0, "PF-BTB-14", 0.86, 0.94, "system",
         "2026-03-14 10:20:00", 1),
        (56, 1014, "1014_pack", 4, 6, "ProductFormsBTB", 1, 1.00, 0.94,
         "ProductCode", 0, "PF-BTB-14", 1.00, 0.94, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 15: TXN 1015 - ProductFormsIFB improvement
        # ---------------------------------------------------------------
        (57, 1015, "1015_pack", 1, 15, "Other", 1, 0.79, 0.92,
         "DocumentType", 0, "Other", 0.79, 0.92, "system",
         "2026-03-14 10:20:00", 1),
        (58, 1015, "1015_pack", 2, 7, "ProductFormsIFB", 1, 1.00, 0.92,
         "DocumentType", 0, "ProductFormsIFB", 1.00, 0.92, "validator",
         "2026-03-14 10:25:00", 1),
        (59, 1015, "1015_pack", 3, 15, "Other", 1, 0.73, 0.92,
         "Comment", 0, "Scanned product form", 0.73, 0.92, "system",
         "2026-03-14 10:20:00", 1),
        (60, 1015, "1015_pack", 4, 7, "ProductFormsIFB", 1, 1.00, 0.92,
         "ProductCode", 0, "PF-IFB-15", 1.00, 0.92, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 16: TXN 1016 - ProductFormsIFB extraction improvement
        # ---------------------------------------------------------------
        (61, 1016, "1016_pack", 1, 7, "ProductFormsIFB", 1, 0.91, 0.95,
         "DocumentType", 0, "ProductFormsIFB", 0.91, 0.95, "system",
         "2026-03-14 10:20:00", 1),
        (62, 1016, "1016_pack", 2, 7, "ProductFormsIFB", 1, 1.00, 0.95,
         "DocumentType", 0, "ProductFormsIFB", 1.00, 0.95, "validator",
         "2026-03-14 10:25:00", 1),
        (63, 1016, "1016_pack", 3, 7, "ProductFormsIFB", 1, 0.00, 0.95,
         "ProductCode", 1, "", 0.00, 0.95, "system",
         "2026-03-14 10:20:00", 1),
        (64, 1016, "1016_pack", 4, 7, "ProductFormsIFB", 1, 1.00, 0.95,
         "ProductCode", 0, "PF-IFB-16", 1.00, 0.95, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 17: TXN 1017 - ProductFormsICIB stable
        # ---------------------------------------------------------------
        (65, 1017, "1017_pack", 1, 8, "ProductFormsICIB", 1, 0.92, 0.96,
         "DocumentType", 0, "ProductFormsICIB", 0.92, 0.96, "system",
         "2026-03-14 10:20:00", 1),
        (66, 1017, "1017_pack", 2, 8, "ProductFormsICIB", 1, 1.00, 0.96,
         "DocumentType", 0, "ProductFormsICIB", 1.00, 0.96, "validator",
         "2026-03-14 10:25:00", 1),
        (67, 1017, "1017_pack", 3, 8, "ProductFormsICIB", 1, 0.89, 0.96,
         "FacilityLimit", 0, "500000", 0.89, 0.96, "system",
         "2026-03-14 10:20:00", 1),
        (68, 1017, "1017_pack", 4, 8, "ProductFormsICIB", 1, 1.00, 0.96,
         "FacilityLimit", 0, "500000", 1.00, 0.96, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 18: TXN 1018 - ProductFormsICIB extraction regression
        # ---------------------------------------------------------------
        (69, 1018, "1018_pack", 1, 8, "ProductFormsICIB", 1, 0.91, 0.95,
         "DocumentType", 0, "ProductFormsICIB", 0.91, 0.95, "system",
         "2026-03-14 10:20:00", 1),
        (70, 1018, "1018_pack", 2, 8, "ProductFormsICIB", 1, 1.00, 0.95,
         "DocumentType", 0, "ProductFormsICIB", 1.00, 0.95, "validator",
         "2026-03-14 10:25:00", 1),
        (71, 1018, "1018_pack", 3, 8, "ProductFormsICIB", 1, 0.88, 0.95,
         "FacilityLimit", 0, "750000", 0.88, 0.95, "system",
         "2026-03-14 10:20:00", 1),
        (72, 1018, "1018_pack", 4, 8, "ProductFormsICIB", 1, 1.00, 0.95,
         "FacilityLimit", 0, "750000", 1.00, 0.95, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 19: TXN 1019 - candidate regression (Passport -> ID)
        # ---------------------------------------------------------------
        (73, 1019, "1019_pack", 1, 4, "Passport", 1, 0.93, 0.97,
         "DocumentType", 0, "Passport", 0.93, 0.97, "system",
         "2026-03-14 10:20:00", 1),
        (74, 1019, "1019_pack", 2, 4, "Passport", 1, 1.00, 0.97,
         "DocumentType", 0, "Passport", 1.00, 0.97, "validator",
         "2026-03-14 10:25:00", 1),
        (75, 1019, "1019_pack", 3, 4, "Passport", 1, 0.90, 0.97,
         "PassportNumber", 0, "C1122334", 0.90, 0.97, "system",
         "2026-03-14 10:20:00", 1),
        (76, 1019, "1019_pack", 4, 4, "Passport", 1, 1.00, 0.97,
         "PassportNumber", 0, "C1122334", 1.00, 0.97, "validator",
         "2026-03-14 10:25:00", 1),

        # ---------------------------------------------------------------
        # Scenario 20: TXN 1020 - identity improvement (Passport -> ID)
        # ---------------------------------------------------------------
        (77, 1020, "1020_pack", 1, 4, "Passport", 1, 0.84, 0.96,
         "DocumentType", 0, "Passport", 0.84, 0.96, "system",
         "2026-03-14 10:20:00", 1),
        (78, 1020, "1020_pack", 2, 5, "IdentityDocument", 1, 1.00, 0.96,
         "DocumentType", 0, "IdentityDocument", 1.00, 0.96, "validator",
         "2026-03-14 10:25:00", 1),
        (79, 1020, "1020_pack", 3, 4, "Passport", 1, 0.81, 0.96,
         "PassportNumber", 0, "C7788990", 0.81, 0.96, "system",
         "2026-03-14 10:20:00", 1),
        (80, 1020, "1020_pack", 4, 5, "IdentityDocument", 1, 1.00, 0.96,
         "IdentityNumber", 0, "9201015009081", 1.00, 0.96, "validator",
         "2026-03-14 10:25:00", 1),
    ]

    _write_sheet(wb, "DocumentData", headers=doc_data_headers, rows=doc_data_rows)

    # --- DocumentData_Candidate ---
    # Candidate output rows mirror DocumentData schema and represent current artifact predictions.
    candidate_rows = [
        (10001, 1001, "1001_pack", 1, 5, "IdentityDocument", 1, 0.94, 0.98, "DocumentType", 0, "IdentityDocument", 0.94, 0.98, "candidate", "2026-03-15 09:00:00", 1),
        (10002, 1001, "1001_pack", 3, 5, "IdentityDocument", 1, 0.93, 0.98, "IdentityNumber", 0, "8001015009087", 0.93, 0.98, "candidate", "2026-03-15 09:00:00", 1),
        (10003, 1002, "1002_pack", 1, 5, "IdentityDocument", 1, 0.89, 0.97, "DocumentType", 0, "IdentityDocument", 0.89, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10004, 1002, "1002_pack", 3, 5, "IdentityDocument", 1, 0.86, 0.97, "IdentityNumber", 0, "9002026009088", 0.86, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10005, 1003, "1003_pack", 1, 15, "Other", 1, 0.90, 0.95, "DocumentType", 0, "Other", 0.90, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10006, 1003, "1003_pack", 3, 15, "Other", 1, 0.84, 0.95, "Comment", 0, "Out of scope brochure", 0.84, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10007, 1004, "1004_pack", 1, 1, "ApplicationForm", 1, 0.94, 0.96, "DocumentType", 0, "ApplicationForm", 0.94, 0.96, "candidate", "2026-03-15 09:00:00", 1),
        (10008, 1004, "1004_pack", 3, 1, "ApplicationForm", 1, 0.88, 0.96, "EntityName", 0, "Acme Holdings Ltd", 0.88, 0.96, "candidate", "2026-03-15 09:00:00", 1),
        (10009, 1005, "1005_pack", 1, 1, "ApplicationForm", 1, 0.87, 0.97, "DocumentType", 0, "ApplicationForm", 0.87, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10010, 1005, "1005_pack", 3, 1, "ApplicationForm", 1, 0.84, 0.97, "ApplicationDate", 0, "2026-01-15", 0.84, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10011, 1006, "1006_pack", 1, 5, "IdentityDocument", 1, 0.92, 0.96, "DocumentType", 0, "IdentityDocument", 0.92, 0.96, "candidate", "2026-03-15 09:00:00", 1),
        (10012, 1006, "1006_pack", 3, 5, "IdentityDocument", 1, 0.91, 0.96, "IdentityNumber", 0, "8503035009089", 0.91, 0.96, "candidate", "2026-03-15 09:00:00", 1),
        (10013, 1007, "1007_pack", 1, 4, "Passport", 1, 0.88, 0.94, "DocumentType", 0, "Passport", 0.88, 0.94, "candidate", "2026-03-15 09:00:00", 1),
        (10014, 1007, "1007_pack", 3, 4, "Passport", 1, 0.84, 0.94, "PassportNumber", 0, "B7654321", 0.84, 0.94, "candidate", "2026-03-15 09:00:00", 1),
        (10015, 1008, "1008_pack", 1, 1, "ApplicationForm", 1, 0.93, 0.97, "DocumentType", 0, "ApplicationForm", 0.93, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10016, 1008, "1008_pack", 3, 1, "ApplicationForm", 1, 0.89, 0.97, "ApplicationDate", 0, "2026-02-20", 0.89, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        # Intentional regression case: high-confidence false positive on out-of-scope.
        (10017, 1009, "1009_pack", 1, 4, "Passport", 1, 0.91, 0.92, "DocumentType", 0, "Passport", 0.91, 0.92, "candidate", "2026-03-15 09:00:00", 1),
        (10018, 1009, "1009_pack", 3, 4, "Passport", 1, 0.87, 0.92, "PassportNumber", 0, "Z0099881", 0.87, 0.92, "candidate", "2026-03-15 09:00:00", 1),
        (10019, 1010, "1010_pack", 1, 4, "Passport", 1, 0.93, 0.97, "DocumentType", 0, "Passport", 0.93, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10020, 1010, "1010_pack", 3, 4, "Passport", 1, 0.90, 0.97, "PassportNumber", 0, "P8899001", 0.90, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10021, 1011, "1011_pack", 1, 3, "Windeed", 1, 0.86, 0.93, "DocumentType", 0, "Windeed", 0.86, 0.93, "candidate", "2026-03-15 09:00:00", 1),
        (10022, 1011, "1011_pack", 3, 3, "Windeed", 1, 0.83, 0.93, "DeedNumber", 0, "WD-556677", 0.83, 0.93, "candidate", "2026-03-15 09:00:00", 1),
        (10023, 1012, "1012_pack", 1, 3, "Windeed", 1, 0.90, 0.94, "DocumentType", 0, "Windeed", 0.90, 0.94, "candidate", "2026-03-15 09:00:00", 1),
        (10024, 1012, "1012_pack", 3, 3, "Windeed", 1, 0.88, 0.94, "OwnerName", 0, "Blue Sky Properties", 0.88, 0.94, "candidate", "2026-03-15 09:00:00", 1),
        (10025, 1013, "1013_pack", 1, 6, "ProductFormsBTB", 1, 0.91, 0.95, "DocumentType", 0, "ProductFormsBTB", 0.91, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10026, 1013, "1013_pack", 3, 6, "ProductFormsBTB", 1, 0.89, 0.95, "ProductCode", 0, "PF-BTB-13", 0.89, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10027, 1014, "1014_pack", 1, 15, "Other", 1, 0.89, 0.94, "DocumentType", 0, "Other", 0.89, 0.94, "candidate", "2026-03-15 09:00:00", 1),
        (10028, 1014, "1014_pack", 3, 15, "Other", 1, 0.85, 0.94, "Comment", 0, "Routed as Other by new prompt", 0.85, 0.94, "candidate", "2026-03-15 09:00:00", 1),
        (10029, 1015, "1015_pack", 1, 7, "ProductFormsIFB", 1, 0.86, 0.92, "DocumentType", 0, "ProductFormsIFB", 0.86, 0.92, "candidate", "2026-03-15 09:00:00", 1),
        (10030, 1015, "1015_pack", 3, 7, "ProductFormsIFB", 1, 0.82, 0.92, "ProductCode", 0, "PF-IFB-15", 0.82, 0.92, "candidate", "2026-03-15 09:00:00", 1),
        (10031, 1016, "1016_pack", 1, 7, "ProductFormsIFB", 1, 0.92, 0.95, "DocumentType", 0, "ProductFormsIFB", 0.92, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10032, 1016, "1016_pack", 3, 7, "ProductFormsIFB", 1, 0.89, 0.95, "ProductCode", 0, "PF-IFB-16", 0.89, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10033, 1017, "1017_pack", 1, 8, "ProductFormsICIB", 1, 0.93, 0.96, "DocumentType", 0, "ProductFormsICIB", 0.93, 0.96, "candidate", "2026-03-15 09:00:00", 1),
        (10034, 1017, "1017_pack", 3, 8, "ProductFormsICIB", 1, 0.90, 0.96, "FacilityLimit", 0, "500000", 0.90, 0.96, "candidate", "2026-03-15 09:00:00", 1),
        (10035, 1018, "1018_pack", 1, 8, "ProductFormsICIB", 1, 0.92, 0.95, "DocumentType", 0, "ProductFormsICIB", 0.92, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10036, 1018, "1018_pack", 3, 8, "ProductFormsICIB", 1, 0.00, 0.95, "FacilityLimit", 1, "", 0.00, 0.95, "candidate", "2026-03-15 09:00:00", 1),
        (10037, 1019, "1019_pack", 1, 5, "IdentityDocument", 1, 0.90, 0.97, "DocumentType", 0, "IdentityDocument", 0.90, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10038, 1019, "1019_pack", 3, 5, "IdentityDocument", 1, 0.87, 0.97, "IdentityNumber", 0, "9101015009080", 0.87, 0.97, "candidate", "2026-03-15 09:00:00", 1),
        (10039, 1020, "1020_pack", 1, 5, "IdentityDocument", 1, 0.88, 0.96, "DocumentType", 0, "IdentityDocument", 0.88, 0.96, "candidate", "2026-03-15 09:00:00", 1),
        (10040, 1020, "1020_pack", 3, 5, "IdentityDocument", 1, 0.84, 0.96, "IdentityNumber", 0, "9201015009081", 0.84, 0.96, "candidate", "2026-03-15 09:00:00", 1),
    ]
    _write_sheet(wb, "DocumentData_Candidate", headers=doc_data_headers, rows=candidate_rows)

    # --- ai.ModelData ---
    model_data_headers = [
        "ModelDataID", "TransactionID", "ModelStageID", "ModelTypeID",
        "ModelNameID", "ModelVersion", "In_Argument_Field",
        "In_Argument_Value", "RecordDateTime", "Active",
    ]

    model_data_rows = [
        (1, 1001, 1, 2, 9001, "bundle_v17", "run_request_json",
         '{"run_id":"RUN-2026-03-14-001"}', "2026-03-14 10:00:00", 1),
        (2, 1001, 2, 2, 9001, "bundle_v17", "change_summary_json",
         '{"prompt_changed":true}', "2026-03-14 10:00:02", 1),
        (3, 1001, 2, 2, 9002, "bundle_v17", "selected_transaction_ids_json",
         "[1001,1002,1003,1004,1005]", "2026-03-14 10:00:04", 1),
        (4, 1001, 2, 2, 9004, "bundle_v17", "regression_findings_json",
         '{"count":2}', "2026-03-14 10:00:06", 1),
        (5, 1001, 2, 2, 9008, "bundle_v17", "root_causes_json",
         '{"primary":"prompt broadened"}', "2026-03-14 10:00:08", 1),
        (6, 1001, 2, 2, 9010, "bundle_v17", "final_run_packet_json",
         '{"verdict":"WARN"}', "2026-03-14 10:00:10", 1),
    ]

    _write_sheet(wb, "ai.ModelData", headers=model_data_headers, rows=model_data_rows)

    # --- ExceptionLogs (headers only) ---
    _write_sheet(
        wb,
        "ExceptionLogs",
        headers=[
            "ExceptionLogID", "TransactionID", "ExceptionType",
            "ExceptionMessage", "ExceptionDateTime",
            "RelatedProcessStageID", "Active",
        ],
        rows=[
            (1, 1005, "ValidationWarning", "Low OCR quality on baseline classification run", "2026-03-14 10:01:00", 1, 1),
            (2, 1008, "MissingFieldWarning", "ApplicationDate missing in baseline extraction", "2026-03-14 10:11:00", 3, 1),
            (3, 1009, "OutOfScopeWarning", "Candidate produced in-scope prediction for out-of-scope doc", "2026-03-15 09:01:00", 1, 1),
            (4, 1011, "ClassificationBaselineMismatch", "Baseline routed Windeed to Other", "2026-03-14 10:21:00", 1, 1),
            (5, 1014, "CandidateRegression", "Candidate routed ProductFormsBTB to Other", "2026-03-15 09:02:00", 1, 1),
            (6, 1018, "CandidateMissingField", "Candidate missing FacilityLimit on ProductFormsICIB", "2026-03-15 09:03:00", 3, 1),
            (7, 1019, "HighConfidenceWrong", "Candidate high-confidence wrong doc type on Passport", "2026-03-15 09:04:00", 1, 1),
        ],
    )

    # --- api.APIData (headers only) ---
    _write_sheet(
        wb,
        "api.APIData",
        headers=[
            "APIDataID", "TransactionID", "APIName", "StatusCode",
            "RequestDateTime", "ResponseDateTime", "ResponseTimeMs", "Active",
        ],
        rows=[
            (1, 1002, "ClassificationAPI", 200, "2026-03-15 09:00:00", "2026-03-15 09:00:01", 830, 1),
            (2, 1005, "ExtractionAPI", 200, "2026-03-15 09:00:02", "2026-03-15 09:00:03", 910, 1),
            (3, 1009, "ClassificationAPI", 200, "2026-03-15 09:00:04", "2026-03-15 09:00:05", 1020, 1),
            (4, 1011, "ClassificationAPI", 200, "2026-03-15 09:00:06", "2026-03-15 09:00:07", 860, 1),
            (5, 1014, "ClassificationAPI", 200, "2026-03-15 09:00:08", "2026-03-15 09:00:09", 940, 1),
            (6, 1018, "ExtractionAPI", 200, "2026-03-15 09:00:10", "2026-03-15 09:00:11", 980, 1),
            (7, 1019, "ClassificationAPI", 200, "2026-03-15 09:00:12", "2026-03-15 09:00:13", 1025, 1),
        ],
    )

    path = os.path.join(output_dir, "DocumentAI_EvidenceStore.xlsx")
    wb.save(path)
    print(f"  Created: {path}")
    return path


# ---------------------------------------------------------------------------
# Workbook 2 — AgenticTesting_AuditLog.xlsx
# ---------------------------------------------------------------------------

def _create_audit_log(output_dir: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(
        wb,
        "RunRegistry",
        headers=[
            "RunID", "ProcessName", "RunMode", "StartDateTime", "EndDateTime",
            "Status", "Verdict", "Confidence", "TransactionsAnalyzed",
            "InitialTransactionCount", "FinalTransactionCount",
            "PromptVersionLabel_Old", "PromptVersionLabel_New",
            "Model_Old", "Model_New", "WorkspaceKey", "TriggeredBy", "Notes",
        ],
    )

    _write_sheet(
        wb,
        "MaestroInput",
        headers=[
            "RunID", "ReceivedDateTime", "InputSection",
            "FieldName", "FieldValue", "ValueType",
        ],
    )

    _write_sheet(
        wb,
        "AgentEvents",
        headers=[
            "RunID", "EventID", "AgentName", "EventType", "Timestamp",
            "DurationSeconds", "InputRef", "OutputRef", "Summary", "Status",
        ],
    )

    _write_sheet(
        wb,
        "ToolCalls",
        headers=[
            "RunID", "ToolCallID", "AgentName", "ToolName", "Timestamp",
            "InputSummary", "RowsRead", "RowsWritten",
            "DurationSeconds", "Status", "Notes",
        ],
    )

    _write_sheet(
        wb,
        "ScopeChanges",
        headers=[
            "RunID", "Timestamp", "Reason",
            "OldScope", "NewScope", "RequestedByAgent",
        ],
    )

    _write_sheet(
        wb,
        "RerunRequests",
        headers=[
            "RunID", "Timestamp", "RequestedByAgent",
            "PatternName", "TransactionIDs", "Reason", "Status",
        ],
    )

    _write_sheet(
        wb,
        "WarningsAndErrors",
        headers=[
            "RunID", "Timestamp", "Source", "Severity",
            "Message", "RelatedTransactionID", "RelatedAgent",
        ],
    )

    _write_sheet(
        wb,
        "OutputArtifacts",
        headers=[
            "RunID", "Timestamp", "ArtifactType",
            "ArtifactURI", "CreatedByAgent", "Status",
        ],
    )

    path = os.path.join(output_dir, "AgenticTesting_AuditLog.xlsx")
    wb.save(path)
    print(f"  Created: {path}")
    return path


# ---------------------------------------------------------------------------
# Workbook 3 — Run_TEMPLATE_Report.xlsx
# ---------------------------------------------------------------------------

def _create_run_template_report(output_dir: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(
        wb,
        "RunSummary",
        headers=[
            "RunID", "StartDateTime", "EndDateTime", "RunMode",
            "Verdict", "Confidence", "TransactionsAnalyzed",
            "PromptVersionOld", "PromptVersionNew",
            "ModelOld", "ModelNew",
        ],
    )

    _write_sheet(
        wb,
        "Scope",
        headers=[
            "RunID", "DateFrom", "DateTo",
            "InitialTransactionCount", "FinalTransactionCount",
            "InitialDocTypes", "ExpandedScope", "ExpansionReason",
        ],
    )

    _write_sheet(
        wb,
        "TransactionsAnalyzed",
        headers=[
            "RunID", "TransactionID", "DocumentID",
            "DocTypeTruth", "DocTypeBaseline", "DocTypeCandidate",
            "ClassificationStatus", "ExtractionStatus",
            "ExceptionPresent", "APIImpactPresent",
        ],
    )

    _write_sheet(
        wb,
        "ClassificationDeltas",
        headers=[
            "RunID", "TransactionID", "DocumentTypeName",
            "BaselineCorrect", "CandidateCorrect", "Changed",
            "ConfidenceBaseline", "ConfidenceCandidate", "Notes",
        ],
    )

    _write_sheet(
        wb,
        "ExtractionDeltas",
        headers=[
            "RunID", "TransactionID", "Field",
            "BaselineValue", "CandidateValue", "TruthValue",
            "BaselineMatch", "CandidateMatch",
            "IsMissingBaseline", "IsMissingCandidate",
        ],
    )

    _write_sheet(
        wb,
        "MissingFieldAnalysis",
        headers=[
            "RunID", "Field", "DocumentTypeName",
            "BaselineMissingRate", "CandidateMissingRate",
            "Delta", "RiskLevel",
        ],
    )

    _write_sheet(
        wb,
        "ExceptionCorrelation",
        headers=[
            "RunID", "TransactionID", "ExceptionType",
            "ExceptionMessage", "RelatedModelFinding", "APIStatusCode",
        ],
    )

    _write_sheet(
        wb,
        "PatchCandidates",
        headers=[
            "RunID", "PatchID", "PatchType", "Target",
            "Description", "Confidence", "RequiresHumanApproval",
            "RecommendedExperiment",
        ],
    )

    _write_sheet(
        wb,
        "FinalRouting",
        headers=[
            "RunID", "BlockRelease", "RequestHumanReview",
            "OpenDefect", "NotifyRoles",
            "ArtifactPDF", "ArtifactHTML", "ArtifactJSON",
        ],
    )

    path = os.path.join(output_dir, "Run_TEMPLATE_Report.xlsx")
    wb.save(path)
    print(f"  Created: {path}")
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create all three Excel workbooks for Document AI Agentic Testing."
    )
    parser.add_argument(
        "--output-dir",
        default="./data",
        metavar="DIR",
        help="Directory to write workbooks into (default: ./data)",
    )
    args = parser.parse_args()

    output_dir = os.path.abspath(args.output_dir)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\nCreating workbooks in: {output_dir}\n")

    evidence_path = _create_evidence_store(output_dir)
    demo_path = os.path.join(output_dir, "DocumentAI_EvidenceStore_Demo.xlsx")
    shutil.copyfile(evidence_path, demo_path)
    print(f"  Created: {demo_path}")
    _create_audit_log(output_dir)
    _create_run_template_report(output_dir)

    print("\nDone. All workbooks created successfully.")


if __name__ == "__main__":
    main()
