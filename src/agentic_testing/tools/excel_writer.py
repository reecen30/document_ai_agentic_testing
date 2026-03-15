"""
Excel Writer tools — write outputs to run workspaces, never to the evidence store.
"""
import json
import os
from typing import Any, Dict, List, Optional, Type
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from crewai.tools import BaseTool
from pydantic import BaseModel, Field


def _ensure_dir(path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)


def _style_header_row(ws, num_cols: int) -> None:
    """Apply header styling to row 1."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


class WriteSheetInput(BaseModel):
    workbook_path: str = Field(..., description="Absolute path to output workbook")
    sheet_name: str = Field(..., description="Sheet name to create or overwrite")
    rows: List[Dict[str, Any]] = Field(..., description="List of row dicts")
    overwrite: bool = Field(default=True)


class WriteSheetTool(BaseTool):
    name: str = "write_sheet"
    description: str = (
        "Write a list of rows (as JSON objects) to a named sheet in an Excel workbook. "
        "Creates the workbook if it does not exist. Creates or overwrites the sheet. "
        "Applies header styling."
    )
    args_schema: Type[BaseModel] = WriteSheetInput

    def _run(
        self,
        workbook_path: str,
        sheet_name: str,
        rows: List[Dict[str, Any]],
        overwrite: bool = True,
    ) -> str:
        try:
            _ensure_dir(workbook_path)
            if os.path.exists(workbook_path):
                wb = load_workbook(workbook_path)
            else:
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            if sheet_name in wb.sheetnames:
                if overwrite:
                    del wb[sheet_name]
                else:
                    ws = wb[sheet_name]
                    if rows:
                        for row in rows:
                            ws.append(list(row.values()))
                    wb.save(workbook_path)
                    return json.dumps({"status": "appended", "rows": len(rows), "sheet": sheet_name})

            ws = wb.create_sheet(title=sheet_name)
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                _style_header_row(ws, len(headers))
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])
                # Auto-width
                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            wb.save(workbook_path)
            return json.dumps({"status": "ok", "rows_written": len(rows), "sheet": sheet_name, "path": workbook_path})
        except Exception as e:
            return json.dumps({"error": str(e), "tool": "write_sheet"})


class AppendRowsInput(BaseModel):
    workbook_path: str
    sheet_name: str
    rows: List[Dict[str, Any]]


class AppendRowsTool(BaseTool):
    name: str = "append_rows"
    description: str = "Append rows to an existing sheet in an Excel workbook. Creates workbook/sheet if needed."
    args_schema: Type[BaseModel] = AppendRowsInput

    def _run(self, workbook_path: str, sheet_name: str, rows: List[Dict[str, Any]]) -> str:
        try:
            _ensure_dir(workbook_path)
            if os.path.exists(workbook_path):
                wb = load_workbook(workbook_path)
            else:
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                if rows:
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    _style_header_row(ws, len(headers))
            else:
                ws = wb[sheet_name]

            for row in rows:
                if ws.max_row == 1 and ws.cell(1, 1).value is None:
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    _style_header_row(ws, len(headers))
                    ws.append([row.get(h, "") for h in headers])
                else:
                    # Infer headers from row 1
                    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    ws.append([row.get(h, "") for h in headers if h])

            wb.save(workbook_path)
            return json.dumps({"status": "ok", "rows_appended": len(rows), "sheet": sheet_name})
        except Exception as e:
            return json.dumps({"error": str(e), "tool": "append_rows"})


class WriteJsonFileInput(BaseModel):
    file_path: str
    data: Dict[str, Any]
    indent: int = Field(default=2)


class WriteJsonFileTool(BaseTool):
    name: str = "write_json_file"
    description: str = "Write a dict as JSON to a file. Creates parent directories as needed."
    args_schema: Type[BaseModel] = WriteJsonFileInput

    def _run(self, file_path: str, data: Dict[str, Any], indent: int = 2) -> str:
        try:
            _ensure_dir(file_path)
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=indent, default=str)
            return json.dumps({"status": "ok", "path": file_path})
        except Exception as e:
            return json.dumps({"error": str(e), "tool": "write_json_file"})


# Convenience instances
write_sheet = WriteSheetTool()
append_rows = AppendRowsTool()
write_json_file = WriteJsonFileTool()
